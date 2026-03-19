"""
Extract P&L data from KKH Excel forecast model into JSON.
Reads sheet "P&L view_AOP" from the forecast workbook and produces
a structured JSON file with line-item data and waterfall bridge datasets.
"""

import json
import math
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
SOURCE_FILE = Path(
    r"C:\Users\Fernando Maskobi\OneDrive\Desktop\KKH AI\Forecast\KKH_FCST_2026_vWIP.xlsx"
)
OUTPUT_FILE = Path(
    r"C:\Users\Fernando Maskobi\OneDrive\Desktop\KKH  analyses\KKH AI\kkh-forecast\src\data\financial-data.json"
)

SHEET_NAME = "P&L view_AOP"

# ---------------------------------------------------------------------------
# Ordered periods & type key mapping
# ---------------------------------------------------------------------------
ORDERED_PERIODS = [
    "Jan", "Feb", "Mar", "Q1",
    "Apr", "May", "Jun", "Q2",
    "Jul", "Aug", "Sep", "Q3",
    "Oct", "Nov", "Dec", "Q4",
    "FY",
]

TYPE_MAP = {
    "Actuals": "actuals",
    "Fcst": "fcst",
    "AOP": "aop",
    "LY": "ly",
}

# ---------------------------------------------------------------------------
# Line-item definitions  (id, label, excel_row, format, isBold)
# ---------------------------------------------------------------------------
LINE_ITEMS = [
    ("gross_booked_sales", "Gross Booked Sales", 7, "$", True),
    ("gbs_shipping", "GBS Shipping", 8, "$", False),
    ("gbs_white_glove", "GBS White Glove", 9, "$", False),
    ("gbs_consolidations", "GBS Consolidations", 10, "$", False),
    ("gbs_assembly", "GBS Assembly", 11, "$", False),
    ("discounts", "Gross Sales Discounts", 12, "$", False),
    ("concessions", "Miscellaneous Concessions", 13, "$", False),
    ("damages", "Damages", 14, "$", False),
    ("cancels", "Cancels", 15, "$", False),
    ("returns", "Returns", 16, "$", False),
    ("net_booked", "Net Booked", 17, "$", True),
    ("rev_rec_pct", "Rev Rec %", 19, "%", False),
    ("gross_shipped_sales", "Gross Shipped Sales", 20, "$", True),
    ("shipping_revenue", "Shipping Revenue", 22, "$", False),
    ("shipping_income_consol", "Shipping Income Consolidations", 23, "$", False),
    ("shipping_income_wg", "Shipping Income WG", 24, "$", False),
    ("assembly_income", "Assembly Income", 25, "$", False),
    ("design_service_fee", "Interior Design Service Fee", 26, "$", False),
    ("other_income", "Other Income", 27, "$", False),
    ("net_shipped_sales", "Net Shipped Sales $", 28, "$", True),
    ("mom_qoq", "MoM/QoQ (%)", 29, "%", False),
    ("gross_cogs", "Gross COGS", 31, "$", False),
    ("gm_dollars", "GM$", 32, "$", True),
    ("gm_percent", "GM%", 33, "%", True),
    ("product_gm_pct", "Product GM %", 35, "%", False),
    ("cogs_interior_design", "COGS Interior Design", 37, "$", False),
    ("marketing_cogs", "Marketing COGS", 38, "$", False),
    ("claims_cogs", "Claims COGS", 39, "$", False),
    ("return_cogs", "Return COGS", 40, "$", False),
    ("cogs_tariff", "COGS Tariff", 41, "$", False),
    ("cogs_total", "COGS", 42, "$", True),
    ("cogs_pct", "COGS%", 43, "%", False),
    ("shipping_cost", "Shipping cost", 45, "$", False),
    ("parcel_shipping", "Parcel Shipping", 46, "$", False),
    ("pallet_upcharge", "Pallet upcharge", 47, "$", False),
    ("drop_ship_charge", "Drop Ship Charge", 48, "$", False),
    ("replacement_cogs", "Replacement COGS", 49, "$", False),
    ("shipping_replacements", "Shipping Replacements", 50, "$", False),
    ("shipping_returns", "Shipping Returns", 51, "$", False),
    ("handling_fees", "Handling Fees", 52, "$", False),
    ("furniture_repair", "Furniture Repair", 53, "$", False),
    ("freight_other", "Freight Other", 54, "$", False),
    ("total_cogs", "Total COGS", 55, "$", True),
    ("total_cogs_pct", "as % of Sales", 56, "%", False),
    ("gross_profit", "Gross Profit", 58, "$", True),
    ("gross_profit_pct", "as % of Sales", 59, "%", False),
    ("channels", "Channels", 61, "$", False),
    ("creative_pr", "Creative + PR", 62, "$", False),
    ("marketing_total", "Marketing", 63, "$", True),
    ("channels_pct_gross", "Channels % Gross Sales", 64, "%", False),
    ("mktg_pct_net", "Mktg % Net Sales", 65, "%", False),
    ("employee_costs", "Employee Costs", 67, "$", False),
    ("distribution_costs", "Distribution Costs", 68, "$", False),
    ("credit_card_fees", "Credit Card Fees", 69, "$", False),
    ("it_costs", "IT Costs", 70, "$", False),
    ("rent_occupancy", "Rent & Occupancy", 71, "$", False),
    ("travel_entertainment", "Travel & Entertainment", 72, "$", False),
    ("professional_fees", "Professional Fees", 73, "$", False),
    ("other_expenses", "Other Expenses", 74, "$", False),
    ("other_ga", "Other G&A", 75, "$", False),
    ("ga_pct", "as % of Sales", 76, "%", False),
    ("total_sga", "Total SG&A", 78, "$", True),
    ("sga_pct", "as % of Net Sales", 79, "%", False),
    ("ebitda", "EBITDA", 81, "$", True),
    ("adjusted_ebitda", "Adjusted EBITDA", 82, "$", False),
    ("ebitda_pct", "as % of Sales", 83, "%", False),
]


def safe_round(val, decimals=4):
    """Return rounded numeric value or None."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        if math.isnan(val) or math.isinf(val):
            return None
        return round(val, decimals)
    return None


def build_column_map(ws):
    """
    Scan row 3 (period) and row 6 (type) to build:
      { (period, type_key): col_index, ... }
    where type_key is one of 'actuals', 'fcst', 'aop', 'ly'.
    """
    col_map = {}  # (period, type_key) -> col (1-based)
    max_col = ws.max_column or 100

    # Read row 3 and row 6 across all columns
    current_period = None
    for col in range(1, max_col + 1):
        period_cell = ws.cell(row=3, column=col).value
        type_cell = ws.cell(row=6, column=col).value

        # Period names sometimes span merged cells; carry forward
        if period_cell is not None:
            p = str(period_cell).strip()
            if p in ORDERED_PERIODS:
                current_period = p

        if current_period and type_cell is not None:
            t = str(type_cell).strip()
            if t in TYPE_MAP:
                key = (current_period, TYPE_MAP[t])
                col_map[key] = col

    return col_map


def get_fy_value(line_data, type_key):
    """Get the FY value for a given type from a line item's data dict."""
    fy = line_data.get("FY")
    if fy is None:
        return 0
    v = fy.get(type_key)
    return v if v is not None else 0


def build_bridges(items_by_id):
    """Compute the three waterfall bridge datasets using FY values."""

    # Helper to look up FY values
    def fy(item_id, type_key):
        item = items_by_id.get(item_id)
        if item is None:
            return 0
        return get_fy_value(item["data"], type_key)

    # --- Sales Bridge: LY Net Shipped -> Fcst Net Shipped ---
    ly_nss = fy("net_shipped_sales", "ly")
    fcst_nss = fy("net_shipped_sales", "fcst")
    shipped_delta = fy("gross_shipped_sales", "fcst") - fy("gross_shipped_sales", "ly")
    shipping_rev_delta = fy("shipping_revenue", "fcst") - fy("shipping_revenue", "ly")
    sales_other = (fcst_nss - ly_nss) - shipped_delta - shipping_rev_delta

    sales_bridge = {
        "title": "Net Shipped Sales Bridge",
        "subtitle": "LY to Forecast (FY)",
        "items": [
            {"name": "LY Net Shipped Sales", "value": round(ly_nss, 4), "isTotal": True},
            {"name": "Shipped Sales", "value": round(shipped_delta, 4), "isTotal": False},
            {"name": "Shipping Revenue", "value": round(shipping_rev_delta, 4), "isTotal": False},
            {"name": "Other", "value": round(sales_other, 4), "isTotal": False},
            {"name": "Fcst Net Shipped Sales", "value": round(fcst_nss, 4), "isTotal": True},
        ],
    }

    # --- Gross Profit Bridge: LY GP -> Fcst GP ---
    ly_gp = fy("gross_profit", "ly")
    fcst_gp = fy("gross_profit", "fcst")
    rev_delta = fy("net_shipped_sales", "fcst") - fy("net_shipped_sales", "ly")
    cogs_delta = -(fy("total_cogs", "fcst") - fy("total_cogs", "ly"))
    gp_other = (fcst_gp - ly_gp) - rev_delta - cogs_delta

    gp_bridge = {
        "title": "Gross Profit Bridge",
        "subtitle": "LY to Forecast (FY)",
        "items": [
            {"name": "LY Gross Profit", "value": round(ly_gp, 4), "isTotal": True},
            {"name": "Revenue", "value": round(rev_delta, 4), "isTotal": False},
            {"name": "COGS", "value": round(cogs_delta, 4), "isTotal": False},
            {"name": "Other", "value": round(gp_other, 4), "isTotal": False},
            {"name": "Fcst Gross Profit", "value": round(fcst_gp, 4), "isTotal": True},
        ],
    }

    # --- EBITDA Bridge: LY EBITDA -> Fcst EBITDA ---
    ly_ebitda = fy("ebitda", "ly")
    fcst_ebitda = fy("ebitda", "fcst")
    gp_delta = fy("gross_profit", "fcst") - fy("gross_profit", "ly")
    mktg_delta = -(fy("marketing_total", "fcst") - fy("marketing_total", "ly"))
    total_delta = fcst_ebitda - ly_ebitda
    ga_delta = total_delta - gp_delta - mktg_delta

    ebitda_bridge = {
        "title": "EBITDA Bridge",
        "subtitle": "LY to Forecast (FY)",
        "items": [
            {"name": "LY EBITDA", "value": round(ly_ebitda, 4), "isTotal": True},
            {"name": "Gross Profit", "value": round(gp_delta, 4), "isTotal": False},
            {"name": "Marketing", "value": round(mktg_delta, 4), "isTotal": False},
            {"name": "G&A", "value": round(ga_delta, 4), "isTotal": False},
            {"name": "Fcst EBITDA", "value": round(fcst_ebitda, 4), "isTotal": True},
        ],
    }

    return {
        "sales": sales_bridge,
        "grossProfit": gp_bridge,
        "ebitda": ebitda_bridge,
    }


def main():
    print(f"Opening workbook: {SOURCE_FILE}")
    wb = openpyxl.load_workbook(str(SOURCE_FILE), data_only=True, read_only=True)
    ws = wb[SHEET_NAME]

    print("Building column map from rows 3 and 6...")
    col_map = build_column_map(ws)
    print(f"  Mapped {len(col_map)} (period, type) combinations")

    # Show which periods have which types
    found_periods = set()
    found_types = set()
    for (p, t) in col_map:
        found_periods.add(p)
        found_types.add(t)
    print(f"  Periods found: {sorted(found_periods, key=lambda x: ORDERED_PERIODS.index(x) if x in ORDERED_PERIODS else 99)}")
    print(f"  Types found: {sorted(found_types)}")

    # Extract line items
    print("Extracting line items...")
    line_items = []
    items_by_id = {}

    for item_id, label, excel_row, fmt, is_bold in LINE_ITEMS:
        data = {}
        for period in ORDERED_PERIODS:
            period_data = {}
            for type_key in ["actuals", "fcst", "aop", "ly"]:
                col = col_map.get((period, type_key))
                if col is not None:
                    val = ws.cell(row=excel_row, column=col).value
                    period_data[type_key] = safe_round(val)
                else:
                    period_data[type_key] = None
            data[period] = period_data

        item = {
            "id": item_id,
            "label": label,
            "row": excel_row,
            "format": fmt,
            "isBold": is_bold,
            "data": data,
        }
        line_items.append(item)
        items_by_id[item_id] = item

    wb.close()

    # Build bridge datasets
    print("Computing bridge datasets...")
    bridges = build_bridges(items_by_id)

    # Assemble output
    output = {
        "metadata": {
            "extractedAt": datetime.now(timezone.utc).isoformat(),
            "sourceFile": SOURCE_FILE.name,
        },
        "periods": ORDERED_PERIODS,
        "lineItems": line_items,
        "bridges": bridges,
    }

    # Write JSON
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_FILE.write_text(json.dumps(output, indent=2, default=str), encoding="utf-8")
    print(f"\nOutput written to: {OUTPUT_FILE}")

    # Summary stats
    total_cells = 0
    non_null_cells = 0
    for item in line_items:
        for period in ORDERED_PERIODS:
            for t in ["actuals", "fcst", "aop", "ly"]:
                total_cells += 1
                if item["data"][period][t] is not None:
                    non_null_cells += 1

    print(f"\n--- Summary ---")
    print(f"  Line items: {len(line_items)}")
    print(f"  Periods: {len(ORDERED_PERIODS)}")
    print(f"  Total data cells: {total_cells}")
    print(f"  Non-null cells: {non_null_cells} ({non_null_cells/total_cells*100:.1f}%)")

    # Print a few key FY values as sanity check
    print(f"\n--- Key FY Values (Fcst) ---")
    for check_id in ["gross_booked_sales", "net_shipped_sales", "gross_profit", "ebitda"]:
        item = items_by_id[check_id]
        fcst_val = item["data"]["FY"]["fcst"]
        print(f"  {item['label']}: {fcst_val}")

    # Print bridge summaries
    print(f"\n--- Bridge Summaries ---")
    for key, bridge in bridges.items():
        print(f"  {bridge['title']}:")
        for b_item in bridge["items"]:
            marker = " (total)" if b_item["isTotal"] else ""
            print(f"    {b_item['name']}: {b_item['value']}{marker}")


if __name__ == "__main__":
    main()
